"""Builds the two spreadsheets in the toolkit: a pricing calculator and an
income/expense tracker. Both ship with live formulas so the buyer edits inputs
only — every derived number recalculates itself."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "out")
BRAND = os.environ.get("KIT_BRAND", "HERERA")

INK = "1F1D1A"
ACCENT = "B0654A"
SOFT = "F7F3EB"
LINE = "DED6C8"
INPUT_BG = "FFFDF7"

title_f = Font(name="Calibri", size=20, bold=True, color=INK)
kicker_f = Font(name="Calibri", size=9, bold=True, color=ACCENT)
head_f = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
label_f = Font(name="Calibri", size=11, color=INK)
note_f = Font(name="Calibri", size=9, italic=True, color="6E6659")
result_f = Font(name="Calibri", size=14, bold=True, color=ACCENT)
bold_f = Font(name="Calibri", size=11, bold=True, color=INK)

head_fill = PatternFill("solid", fgColor=INK)
soft_fill = PatternFill("solid", fgColor=SOFT)
input_fill = PatternFill("solid", fgColor=INPUT_BG)
thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '#,##0.00'
PCT = '0%'


def header(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = title_f
    ws["A2"] = subtitle
    ws["A2"].font = note_f
    ws["A3"] = f"{BRAND}  ·  SMALL BUSINESS TOOLKIT"
    ws["A3"].font = kicker_f
    ws.row_dimensions[1].height = 28


def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text.upper())
    c.font = head_f
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = head_fill
        ws.cell(row=row, column=col).font = head_f
    ws.row_dimensions[row].height = 18


def field(ws, row, label, value=None, fmt=None, note=None, editable=True):
    ws.cell(row=row, column=1, value=label).font = label_f
    c = ws.cell(row=row, column=2, value=value)
    c.border = box
    if editable:
        c.fill = input_fill
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = note_f
    return c


# --------------------------------------------------------------- calculator
def pricing_calculator():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Calculator"
    ws.sheet_view.showGridLines = False

    header(ws, "Pricing Calculator",
           "Fill in only the shaded cells. Everything else calculates itself.")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 4

    section(ws, 5, "1 · What you need to earn")
    field(ws, 6, "Take-home pay you want per year", 60000, MONEY,
          "Be honest. This is salary, not revenue.")
    field(ws, 7, "Business costs per year", 6000, MONEY,
          "Software, fees, equipment, marketing, accountant.")
    field(ws, 8, "Tax set-aside", 0.25, PCT,
          "Check your local rate. Guessing low here is why people run out of money.")
    r = field(ws, 9, "Revenue you actually need", "=(B6+B7)/(1-B8)", MONEY,
              "What the business must bill before tax.", editable=False)
    r.font = bold_f
    ws.cell(row=9, column=2).fill = soft_fill

    section(ws, 11, "2 · How much you can actually sell")
    field(ws, 12, "Working weeks per year", 46, None,
          "52 minus holiday and sick time. Nobody works 52.")
    field(ws, 13, "Working hours per week", 40, None)
    field(ws, 14, "% of hours that are billable", 0.6, PCT,
          "Admin, sales and marketing are not billable. 60% is realistic; 80% is a fantasy.")
    b = field(ws, 15, "Billable hours per year", "=B12*B13*B14", '#,##0',
              "The hours you can actually invoice.", editable=False)
    b.font = bold_f
    ws.cell(row=15, column=2).fill = soft_fill

    section(ws, 17, "3 · Your number")
    h = field(ws, 18, "Minimum hourly rate", "=B9/B15", MONEY,
              "Below this you are paying to work.", editable=False)
    h.font = result_f
    ws.cell(row=18, column=2).fill = soft_fill
    d = field(ws, 19, "Minimum day rate", "=B18*B13/5", MONEY,
              "Round up to something memorable.", editable=False)
    d.font = result_f
    ws.cell(row=19, column=2).fill = soft_fill

    section(ws, 21, "4 · Price a specific project")
    field(ws, 22, "Hours you estimate", 20, '#,##0')
    field(ws, 23, "Add a buffer for the unexpected", 0.2, PCT,
          "Every project runs over. 20% is the minimum.")
    field(ws, 24, "Direct costs (stock, licences, subcontractors)", 0, MONEY)
    field(ws, 25, "Profit margin on top", 0.2, PCT,
          "Profit is separate from your salary. It is what lets the business grow.")
    q = field(ws, 26, "Quote this project at", "=((B22*(1+B23)*B18)+B24)*(1+B25)", MONEY,
              "This is your price. Do not discount it in your head before you say it.",
              editable=False)
    q.font = result_f
    ws.cell(row=26, column=2).fill = soft_fill
    field(ws, 27, "Deposit to invoice up front (50%)", "=B26*0.5", MONEY,
          "Work starts when this clears. Not before.", editable=False)
    ws.cell(row=27, column=2).fill = soft_fill

    ws["A29"] = ("If your current rate is well below row 18, that is not a sign you are bad at this — "
                 "it is a sign nobody ever showed you the arithmetic.")
    ws["A29"].font = note_f

    wb.save(os.path.join(OUT, "pricing-calculator.xlsx"))
    print("  ✓ pricing-calculator.xlsx")


# ------------------------------------------------------------------ tracker
def income_tracker():
    wb = Workbook()

    # ---- Income
    ws = wb.active
    ws.title = "Income"
    ws.sheet_view.showGridLines = False
    header(ws, "Income", "Log every payment the day it lands. One row per payment.")

    cols = [("Date", 14), ("Client", 26), ("Description", 34),
            ("Invoice #", 13), ("Amount", 14), ("Paid?", 10)]
    for i, (name, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=5, column=i, value=name.upper())
        c.font = head_f
        c.fill = head_fill
        c.alignment = Alignment(horizontal="left")

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for r in range(6, 206):
        for i in range(1, 7):
            cell = ws.cell(row=r, column=i)
            cell.border = box
            cell.fill = input_fill
        ws.cell(row=r, column=1).number_format = 'dd/mm/yyyy'
        ws.cell(row=r, column=5).number_format = MONEY
        dv.add(ws.cell(row=r, column=6))

    # ---- Expenses
    we = wb.create_sheet("Expenses")
    we.sheet_view.showGridLines = False
    header(we, "Expenses", "Anything you spent to run the business. Keep the receipt.")

    ecols = [("Date", 14), ("Supplier", 26), ("Category", 22),
             ("Description", 30), ("Amount", 14)]
    for i, (name, width) in enumerate(ecols, start=1):
        we.column_dimensions[get_column_letter(i)].width = width
        c = we.cell(row=5, column=i, value=name.upper())
        c.font = head_f
        c.fill = head_fill

    CATS = "Software,Equipment,Marketing,Fees & charges,Subcontractors,Travel,Education,Other"
    dvc = DataValidation(type="list", formula1=f'"{CATS}"', allow_blank=True)
    we.add_data_validation(dvc)
    for r in range(6, 206):
        for i in range(1, 6):
            cell = we.cell(row=r, column=i)
            cell.border = box
            cell.fill = input_fill
        we.cell(row=r, column=1).number_format = 'dd/mm/yyyy'
        we.cell(row=r, column=5).number_format = MONEY
        dvc.add(we.cell(row=r, column=3))

    # ---- Summary
    wsm = wb.create_sheet("Summary", 0)
    wsm.sheet_view.showGridLines = False
    header(wsm, "Year Summary", "This sheet fills itself in from Income and Expenses.")
    wsm.column_dimensions["A"].width = 16
    for col in "BCD":
        wsm.column_dimensions[col].width = 18
    wsm.column_dimensions["E"].width = 46

    section(wsm, 5, "Month by month")
    for i, name in enumerate(["Month", "Income", "Expenses", "Profit"]):
        c = wsm.cell(row=6, column=i + 1, value=name.upper())
        c.font = head_f
        c.fill = head_fill

    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    for i, m in enumerate(months):
        r = 7 + i
        wsm.cell(row=r, column=1, value=m).font = label_f
        # month index is derived from the row so the formulas stay copyable
        wsm.cell(row=r, column=2,
                 value=f'=SUMPRODUCT((MONTH(Income!$A$6:$A$205)={i+1})*(Income!$E$6:$E$205))')
        wsm.cell(row=r, column=3,
                 value=f'=SUMPRODUCT((MONTH(Expenses!$A$6:$A$205)={i+1})*(Expenses!$E$6:$E$205))')
        wsm.cell(row=r, column=4, value=f"=B{r}-C{r}")
        for col in range(2, 5):
            cell = wsm.cell(row=r, column=col)
            cell.number_format = MONEY
            cell.border = box

    tr = 7 + len(months)
    wsm.cell(row=tr, column=1, value="TOTAL").font = bold_f
    for col, letter in ((2, "B"), (3, "C"), (4, "D")):
        c = wsm.cell(row=tr, column=col, value=f"=SUM({letter}7:{letter}{tr-1})")
        c.number_format = MONEY
        c.font = bold_f
        c.fill = soft_fill

    section(wsm, tr + 2, "The numbers that matter")
    rows = [
        ("Total revenue", f"=B{tr}", MONEY, "Everything you billed and were paid."),
        ("Total expenses", f"=C{tr}", MONEY, "Everything it cost to earn that."),
        ("Profit", f"=D{tr}", MONEY, "What the business actually made."),
        ("Profit margin", f"=IF(B{tr}=0,0,D{tr}/B{tr})", PCT,
         "Under 30% usually means you are underpricing, not overspending."),
        ("Average month", f"=D{tr}/12", MONEY, "Your real monthly income."),
        ("Best month", f"=MAX(D7:D{tr-1})", MONEY, "Work out what you did differently. Repeat it."),
    ]
    for i, (label, formula, fmt, note) in enumerate(rows):
        r = tr + 3 + i
        wsm.cell(row=r, column=1, value=label).font = label_f
        c = wsm.cell(row=r, column=2, value=formula)
        c.number_format = fmt
        c.font = result_f if label in ("Profit", "Profit margin") else bold_f
        c.fill = soft_fill
        c.border = box
        wsm.cell(row=r, column=5, value=note).font = note_f

    wb.save(os.path.join(OUT, "income-expense-tracker.xlsx"))
    print("  ✓ income-expense-tracker.xlsx")


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    pricing_calculator()
    income_tracker()
    print("\n2 spreadsheets written to kit/product/out/")
